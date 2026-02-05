"""Report generation for IO Crosscheck — XLSX and HTML outputs."""
from __future__ import annotations

from pathlib import Path
from typing import Sequence

from io_crosscheck.models import MatchResult, Classification, Confidence


# ---------------------------------------------------------------------------
# XLSX Report
# ---------------------------------------------------------------------------

_COLOR_MAP = {
    Classification.BOTH: "92D050",          # green
    Classification.RACK_ONLY: "FFFF00",       # yellow
    Classification.IO_LIST_ONLY: "FF0000",   # red
    Classification.PLC_ONLY: "5B9BD5",       # blue
    Classification.CONFLICT: "FFC000",       # orange
    Classification.SPARE: "D9D9D9",          # grey
}


def generate_xlsx_report(
    results: Sequence[MatchResult],
    output_path: Path,
    summary: dict | None = None,
) -> Path:
    """Write a verification report to XLSX with conditional formatting.

    Returns the path to the written file.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # ---- Detail sheet ----
    ws = wb.active
    ws.title = "Verification Detail"

    headers = [
        "Device Tag", "IO Tag", "Panel", "Rack", "Slot", "Channel",
        "PLC Address", "Module Type", "Classification", "Strategy",
        "Confidence", "PLC Tag Name", "PLC Description", "Conflict",
        "Audit Trail",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, r in enumerate(results, start=2):
        io = r.io_device
        plc = r.plc_tag
        values = [
            io.device_tag if io else "",
            io.io_tag if io else "",
            io.panel if io else "",
            io.rack if io else "",
            io.slot if io else "",
            io.channel if io else "",
            io.plc_address if io else "",
            io.module_type if io else "",
            r.classification.value,
            r.strategy_id if r.strategy_id else "",
            r.confidence.value if r.strategy_id else "",
            plc.name if plc else "",
            plc.description if plc else "",
            "YES" if r.conflict_flag else "",
            " | ".join(r.audit_trail),
        ]
        fill_color = _COLOR_MAP.get(r.classification)
        row_fill = PatternFill(
            start_color=fill_color, end_color=fill_color, fill_type="solid"
        ) if fill_color else None

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx == 9 and row_fill:  # Classification column
                cell.fill = row_fill
                cell.font = Font(bold=True)

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # ---- Summary sheet ----
    ws_sum = wb.create_sheet("Summary")
    ws_sum.cell(row=1, column=1, value="IO Crosscheck — Summary Report").font = Font(bold=True, size=14)

    if summary is None:
        summary = _build_summary(results)

    row = 3
    ws_sum.cell(row=row, column=1, value="Classification").font = Font(bold=True)
    ws_sum.cell(row=row, column=2, value="Count").font = Font(bold=True)
    ws_sum.cell(row=row, column=3, value="Percentage").font = Font(bold=True)
    row += 1

    total = summary.get("total", 0)
    for cls_name, count in summary.get("by_classification", {}).items():
        ws_sum.cell(row=row, column=1, value=cls_name)
        ws_sum.cell(row=row, column=2, value=count)
        pct = f"{count / total * 100:.1f}%" if total > 0 else "0%"
        ws_sum.cell(row=row, column=3, value=pct)
        row += 1

    row += 1
    ws_sum.cell(row=row, column=1, value="Total Devices").font = Font(bold=True)
    ws_sum.cell(row=row, column=2, value=total)

    row += 2
    ws_sum.cell(row=row, column=1, value="Conflicts Requiring Review").font = Font(bold=True)
    ws_sum.cell(row=row, column=2, value=summary.get("conflicts", 0))

    # ---- Conflict sheet ----
    conflicts = [r for r in results if r.conflict_flag]
    if conflicts:
        ws_conf = wb.create_sheet("Conflicts")
        conf_headers = [
            "Device Tag", "IO Tag", "PLC Address", "IO Description",
            "PLC Description", "Audit Trail",
        ]
        for col_idx, h in enumerate(conf_headers, start=1):
            cell = ws_conf.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

        for row_idx, r in enumerate(conflicts, start=2):
            io = r.io_device
            plc = r.plc_tag
            vals = [
                io.device_tag if io else "",
                io.io_tag if io else "",
                io.plc_address if io else "",
                io.device_tag if io else "",
                plc.description if plc else "",
                " | ".join(r.audit_trail),
            ]
            for col_idx, val in enumerate(vals, start=1):
                ws_conf.cell(row=row_idx, column=col_idx, value=val)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path


# ---------------------------------------------------------------------------
# HTML Report
# ---------------------------------------------------------------------------

_HTML_TEMPLATE = """\
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>IO Crosscheck Report</title>
<link rel="stylesheet" href="https://cdn.datatables.net/1.13.6/css/jquery.dataTables.min.css">
<style>
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; margin: 20px; background: #f5f5f5; }}
  h1 {{ color: #2c3e50; }}
  .summary {{ display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 24px; }}
  .card {{ background: white; border-radius: 8px; padding: 16px 24px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); min-width: 140px; }}
  .card .count {{ font-size: 2em; font-weight: bold; }}
  .card .label {{ color: #666; font-size: 0.9em; }}
  .both {{ color: #27ae60; }}
  .both-rack {{ color: #f39c12; }}
  .io-only {{ color: #e74c3c; }}
  .plc-only {{ color: #3498db; }}
  .conflict {{ color: #e67e22; }}
  .spare {{ color: #95a5a6; }}
  table {{ width: 100%; }}
  table.dataTable tbody tr:hover {{ background-color: #eef; }}
  .cls-Both {{ background-color: #d4edda !important; }}
  .cls-Rack-Only {{ background-color: #fff3cd !important; }}
  .cls-IO-List-Only {{ background-color: #f8d7da !important; }}
  .cls-PLC-Only {{ background-color: #d1ecf1 !important; }}
  .cls-Conflict {{ background-color: #ffeaa7 !important; }}
  .cls-Spare {{ background-color: #e9ecef !important; }}
</style>
</head>
<body>
<h1>IO Crosscheck — Verification Report</h1>

<div class="summary">
  <div class="card"><div class="count">{total}</div><div class="label">Total</div></div>
  <div class="card"><div class="count both">{both}</div><div class="label">Both</div></div>
  <div class="card"><div class="count both-rack">{rack_only}</div><div class="label">Rack Only</div></div>
  <div class="card"><div class="count io-only">{io_only}</div><div class="label">IO List Only</div></div>
  <div class="card"><div class="count plc-only">{plc_only}</div><div class="label">PLC Only</div></div>
  <div class="card"><div class="count conflict">{conflicts}</div><div class="label">Conflicts</div></div>
  <div class="card"><div class="count spare">{spares}</div><div class="label">Spares</div></div>
</div>

<table id="results" class="display compact">
<thead>
<tr>
  <th>Device Tag</th><th>IO Tag</th><th>Panel</th><th>Rack</th>
  <th>Slot</th><th>Channel</th><th>PLC Address</th><th>Module Type</th>
  <th>Classification</th><th>Strategy</th><th>Confidence</th>
  <th>PLC Tag</th><th>PLC Description</th><th>Conflict</th>
</tr>
</thead>
<tbody>
{rows}
</tbody>
</table>

<script src="https://code.jquery.com/jquery-3.7.0.min.js"></script>
<script src="https://cdn.datatables.net/1.13.6/js/jquery.dataTables.min.js"></script>
<script>
$(document).ready(function() {{
  $('#results').DataTable({{
    pageLength: 50,
    order: [[8, 'asc']],
  }});
}});
</script>
</body>
</html>
"""


def _cls_css(classification: Classification) -> str:
    return "cls-" + classification.value.replace(" ", "-").replace("(", "").replace(")", "")


def generate_html_report(
    results: Sequence[MatchResult],
    output_path: Path,
) -> Path:
    """Write an interactive HTML verification report."""
    summary = _build_summary(results)

    rows_html = []
    for r in results:
        io = r.io_device
        plc = r.plc_tag
        css = _cls_css(r.classification)
        cols = [
            io.device_tag if io else "",
            io.io_tag if io else "",
            io.panel if io else "",
            io.rack if io else "",
            io.slot if io else "",
            io.channel if io else "",
            io.plc_address if io else "",
            io.module_type if io else "",
            r.classification.value,
            str(r.strategy_id) if r.strategy_id else "",
            r.confidence.value if r.strategy_id else "",
            plc.name if plc else "",
            plc.description if plc else "",
            "YES" if r.conflict_flag else "",
        ]
        tds = "".join(f"<td>{_esc(c)}</td>" for c in cols)
        rows_html.append(f'<tr class="{css}">{tds}</tr>')

    by_cls = summary["by_classification"]
    html = _HTML_TEMPLATE.format(
        total=summary["total"],
        both=by_cls.get(Classification.BOTH.value, 0),
        rack_only=by_cls.get(Classification.RACK_ONLY.value, 0),
        io_only=by_cls.get(Classification.IO_LIST_ONLY.value, 0),
        plc_only=by_cls.get(Classification.PLC_ONLY.value, 0),
        conflicts=summary["conflicts"],
        spares=by_cls.get(Classification.SPARE.value, 0),
        rows="\n".join(rows_html),
    )

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(html, encoding="utf-8")
    return output_path


def _esc(text: str) -> str:
    """Escape HTML special characters."""
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


# ---------------------------------------------------------------------------
# Summary builder
# ---------------------------------------------------------------------------

def _build_summary(results: Sequence[MatchResult]) -> dict:
    by_cls: dict[str, int] = {}
    conflicts = 0
    for r in results:
        key = r.classification.value
        by_cls[key] = by_cls.get(key, 0) + 1
        if r.conflict_flag:
            conflicts += 1

    return {
        "total": len(results),
        "by_classification": by_cls,
        "conflicts": conflicts,
    }
