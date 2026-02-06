"""Input parsers for PLC CSV tag exports and IO List XLSX files."""
from __future__ import annotations

import csv
import re
from pathlib import Path

from io_crosscheck.models import PLCTag, IODevice, RecordType, AddressFormat
from io_crosscheck.normalizers import detect_address_format

_RECORD_TYPES = {
    "TAG": RecordType.TAG,
    "COMMENT": RecordType.COMMENT,
    "ALIAS": RecordType.ALIAS,
    "RCOMMENT": RecordType.RCOMMENT,
}

_BASE_NAME_SUFFIX_RE = re.compile(r":[IOCS]\d*$", re.IGNORECASE)


def _extract_base_name(name: str) -> str:
    """Strip trailing :I, :O, :C, :S, :I1, :O1 etc. from a tag name."""
    return _BASE_NAME_SUFFIX_RE.sub("", name.strip())


def parse_plc_csv(filepath: Path, encoding: str = "latin-1") -> list[PLCTag]:
    """Parse an RSLogix 5000 CSV tag export file.

    Handles TAG, COMMENT, ALIAS, and RCOMMENT record types.
    The RSLogix CSV is non-standard with mixed record types and multi-line descriptions.
    """
    tags: list[PLCTag] = []
    filepath = Path(filepath)

    with open(filepath, "r", encoding=encoding, errors="replace") as f:
        reader = csv.reader(f)
        header = None
        for line_num, row in enumerate(reader, start=1):
            if not row:
                continue
            # Detect header row
            if header is None:
                if row[0].strip().upper() == "TYPE":
                    header = [c.strip().upper() for c in row]
                continue

            record_type_str = row[0].strip().upper()
            if record_type_str not in _RECORD_TYPES:
                continue

            record_type = _RECORD_TYPES[record_type_str]

            # Skip RCOMMENT records — they are rung comments, not tag data
            if record_type == RecordType.RCOMMENT:
                continue

            # Map columns by header position
            def col(name: str) -> str:
                try:
                    idx = header.index(name)
                    return row[idx].strip() if idx < len(row) else ""
                except (ValueError, IndexError):
                    return ""

            name = col("NAME")
            tag = PLCTag(
                record_type=record_type,
                name=name,
                base_name=_extract_base_name(name),
                description=col("DESCRIPTION"),
                datatype=col("DATATYPE"),
                scope=col("SCOPE"),
                specifier=col("SPECIFIER"),
                source_line=line_num,
            )
            tags.append(tag)

    return tags


def parse_io_list_xlsx(filepath: Path, sheet_name: str = "ESCO List") -> list[IODevice]:
    """Parse an IO List XLSX file from the specified sheet.

    Reads panel, rack, group, slot, channel, PLC IO address, IO tag,
    device tag, module type, module, and range data.
    """
    import openpyxl

    filepath = Path(filepath)
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]

    devices: list[IODevice] = []
    header: list[str] | None = None
    header_map: dict[str, int] = {}

    for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [str(c).strip() if c is not None else "" for c in row]

        # Detect header row
        if header is None:
            if any("panel" in c.lower() for c in cells):
                header = [c.lower() for c in cells]
                for i, h in enumerate(header):
                    header_map[h] = i
            continue

        # Skip empty rows
        if all(c == "" for c in cells):
            continue

        # Skip rows with no meaningful IO data (no device tag, IO tag, or PLC address)
        def _peek(name: str) -> str:
            idx = header_map.get(name.lower())
            return cells[idx] if idx is not None and idx < len(cells) else ""

        if not _peek("device tag") and not _peek("io tag") and not _peek("plc io address"):
            continue

        def col(name: str) -> str:
            name_l = name.lower()
            idx = header_map.get(name_l)
            if idx is not None and idx < len(cells):
                return cells[idx]
            return ""

        plc_address = col("plc io address")
        fmt_str = detect_address_format(plc_address)
        if fmt_str == "CLX":
            addr_fmt = AddressFormat.CLX
        elif fmt_str == "PLC5":
            addr_fmt = AddressFormat.PLC5
        else:
            addr_fmt = AddressFormat.UNKNOWN

        device = IODevice(
            panel=col("panel"),
            rack=col("rack"),
            group=col("group"),
            slot=col("slot"),
            channel=col("channel"),
            plc_address=plc_address,
            io_tag=col("io tag"),
            device_tag=col("device tag"),
            module_type=col("module type"),
            module=col("module"),
            range_low=col("range low"),
            range_high=col("range high"),
            units=col("units"),
            address_format=addr_fmt,
            source_row=row_num,
        )
        devices.append(device)

    wb.close()
    return devices


def parse_rack_layouts(filepath: Path, sheet_name: str = "Rack Layouts") -> dict:
    """Parse the Rack Layouts sheet for physical slot-to-device cross-reference."""
    import openpyxl

    filepath = Path(filepath)
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]

    layouts: dict[str, str] = {}
    header: list[str] | None = None

    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if header is None:
            header = [c.lower() for c in cells]
            continue
        # Build a key from available location columns and map to device
        if len(cells) >= 2:
            key = "|".join(cells[:-1]).lower()
            device = cells[-1]
            if device:
                layouts[key] = device

    wb.close()
    return layouts
